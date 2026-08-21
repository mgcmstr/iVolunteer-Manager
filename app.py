# -*- coding: utf-8 -*-
"""i志愿 考勤比对工具 - FastAPI 后端

启动：venv\\Scripts\\python app.py  →  http://localhost:8000
"""
import json
import logging
import os
import threading
import time
import traceback

from fastapi import FastAPI, HTTPException
from fastapi.responses import FileResponse, JSONResponse
from fastapi.staticfiles import StaticFiles

import compare
import login
from api_client import GdzyzClient, GdzyzError, dedup_records, load_debug_last

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
EXPORT_DIR = os.path.join(BASE_DIR, "exports")
os.makedirs(EXPORT_DIR, exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
)
logger = logging.getLogger("izy")

app = FastAPI(title="i志愿 考勤比对工具")

# ---------- 客户端工具 ----------
_client_lock = threading.Lock()
_client = None


def get_client() -> GdzyzClient:
    global _client
    token = login.load_token()
    if not token:
        raise HTTPException(status_code=401, detail="未登录：请先登录或粘贴 Token")
    with _client_lock:
        if _client is None or _client.token != token:
            _client = GdzyzClient(token)
        return _client


# ---------- 活动字段尽力提取 ----------
def _try_keys(item: dict, keys):
    for k in keys:
        if k in item and item[k] not in (None, "", [], {}):
            return item[k]
    return None


def _extract_activity(item: dict) -> dict:
    """从活动记录中尽力提取展示字段（字段名已联调校准）

    实测字段：missionId / subject / state / createDatetime / finishDatetime / missionType
    """
    aid = _try_keys(item, ["missionId", "id", "mid", "activityId"])
    name = _try_keys(item, ["subject", "missionName", "name", "activityName", "title"])
    start_time = _try_keys(item, ["startDate", "startTime", "createDatetime", "beginTime"])
    end_time = _try_keys(item, ["endDate", "finishDatetime", "endTime"])
    status = _try_keys(item, ["state", "status", "missionStatus", "auditStatus"])
    org = _try_keys(item, ["districtName", "orgName", "organizationName", "groupName"])
    contact = _try_keys(item, ["contactName", "linkman", "contactPerson"])
    phone = _try_keys(item, ["contactPhone", "linkPhone", "phone"])
    return {
        "id": aid,
        "name": name,
        "startTime": start_time,
        "endTime": end_time,
        "status": status,
        "orgName": org,
        "contactName": contact,
        "contactPhone": phone,
        "raw": item,
    }


# ---------- 状态 ----------
@app.get("/api/status")
def api_status():
    token = login.load_token()
    client = None
    valid = False
    if token:
        try:
            client = GdzyzClient(token)
            valid = client.check_login()
        except Exception:
            valid = False
    return {
        "logged_in": bool(token),
        "token_valid": valid,
        "has_token": bool(token),
    }


# ---------- 登录 ----------
@app.post("/api/login/token")
async def api_login_token(req: dict):
    token = (req.get("token") or "").strip()
    if not token:
        raise HTTPException(status_code=400, detail="Token 不能为空")
    login.set_token_manual(token)
    return {"ok": True, "message": "Token 已保存"}


@app.post("/api/login/clear")
def api_login_clear():
    login.clear_token()
    return {"ok": True, "message": "Token 已清除"}


# ---------- 活动查询 ----------
@app.get("/api/activities")
def api_activities(limit: int = 10):
    client = get_client()
    try:
        result = client.get_activities()
    except GdzyzError as e:
        raise HTTPException(status_code=502, detail=str(e))
    items = result.get("list") or []
    # 按活动开始时间降序（最近开始的在前）
    def sort_key(it):
        a = _extract_activity(it)
        return str(a["startTime"] or "")
    items_sorted = sorted(items, key=sort_key, reverse=True)
    activities = [_extract_activity(it) for it in items_sorted[:limit]]
    return {
        "activities": activities,
        "total": result.get("total"),
        "returned": len(activities),
        "warning": "接口字段尚未校准，如有显示异常请打开调试面板查看原始数据" if not activities else None,
    }


# ---------- 活动详情（名单+考勤+比对） ----------
@app.get("/api/activity/{mission_id}/detail")
def api_activity_detail(mission_id: str, timeout: int = 120):
    client = get_client()
    try:
        attendees_raw = client.get_all_personal(mission_id, max_pages=100)
        attendance_raw = client.get_all_attendance(mission_id, max_pages=100)
    except GdzyzError as e:
        raise HTTPException(status_code=502, detail=str(e))

    # 去重：防止接口重复返回同一人导致人数虚高
    attendees, dup_att = dedup_records(attendees_raw)
    attendance, dup_atd = dedup_records(
        attendance_raw, id_keys=("missionServiceLogId", "userid", "userName"))

    result = compare.compare(attendees, attendance)
    stats = dict(result["stats"])
    stats["duplicates_removed_attendees"] = dup_att
    stats["duplicates_removed_attendance"] = dup_atd
    warnings = list(result["warnings"])
    if dup_att > 0:
        warnings.append(f"名单接口返回了 {dup_att} 条重复记录，已自动去重（实际参加 {len(attendees)} 人）")
    if dup_atd > 0:
        warnings.append(f"考勤接口返回了 {dup_atd} 条重复记录，已自动去重")
    return {
        "mission_id": mission_id,
        "attendees": attendees,
        "attendance": attendance,
        "signed": result["signed"],
        "signed_both": result["signed_both"],
        "signed_no_checkout": result["signed_no_checkout"],
        "not_signed": result["not_signed"],
        "anomalies": result["anomalies"],
        "stats": stats,
        "warnings": warnings,
    }


# ---------- 导出 Excel ----------
@app.get("/api/activity/{mission_id}/export")
def api_activity_export(mission_id: str):
    client = get_client()
    try:
        attendees_raw = client.get_all_personal(mission_id, max_pages=100)
        attendance_raw = client.get_all_attendance(mission_id, max_pages=100)
    except GdzyzError as e:
        raise HTTPException(status_code=502, detail=str(e))

    attendees, _ = dedup_records(attendees_raw)
    attendance, _ = dedup_records(
        attendance_raw, id_keys=("missionServiceLogId", "userid", "userName"))

    result = compare.compare(attendees, attendance)
    return _build_excel_response(mission_id, result)


def _build_excel_response(mission_id: str, result: dict):
    """生成 Excel：已签到 / 未签到 / 异常 三个 sheet"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()

    # 人员名单行（附考勤备注）
    def person_row(rec, prefix=""):
        p = rec.get("person") if isinstance(rec, dict) else rec
        row = []
        if isinstance(p, dict):
            name = compare._pick(p, compare.NAME_KEYS)
            phone = compare._pick(p, compare.PHONE_KEYS)
            # 身份证列：优先真实证件号字段，避免误取 userid
            idv = compare._pick(p, ["idcardCode", "idCard", "idcard", "cardNo",
                                    "certificateNo", "identityCard", "IDCard"])
            gender = compare._pick(p, ["gender", "sex"])
            gender_txt = "男" if gender in (1, "1", "M", "male") else ("女" if gender in (2, "2", "F", "female") else gender)
            row = [name or "", gender_txt or "", phone or "", idv or "",
                   compare._pick(p, ["signDatetime", "enrollTime", "signupTime",
                                     "applyTime", "inviteTime", "addtime"]),
                   compare._pick(p, ["selectedDatetime", "hireTime", "employTime"]),
                   compare._pick(p, compare.STATUS_KEYS)]
        # 备注：签到/签退状态 + 匹配方式
        notes = []
        if isinstance(rec, dict):
            att = rec.get("attendance_record")
            if att:
                if compare._has_checkout(att):
                    notes.append("已签到+已签退")
                else:
                    notes.append("已签到-未签退")
            else:
                notes.append("未签到")
            if rec.get("match_method"):
                notes.append(f"匹配:{rec['match_method']}")
        row.append("、".join(notes))
        return row

    def write_sheet(ws, title, rows, fill_hex=None):
        ws.title = title
        headers = ["姓名", "性别", "手机号", "身份证号", "报名时间", "录用时间", "状态", "备注"]
        ws.append(headers)
        for c in ws[1]:
            c.font = Font(bold=True)
            c.fill = PatternFill("solid", fgColor="D9E1F2")
            c.alignment = Alignment(horizontal="center")
        for rec in rows:
            ws.append(person_row(rec))
        if fill_hex:
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
                for c in row:
                    c.fill = PatternFill("solid", fgColor=fill_hex)
        for col in ws.columns:
            max_len = max(len(str(c.value or "")) for c in col[:200])
            ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 8), 40)
        ws.freeze_panes = "A2"

    ws1 = wb.active
    write_sheet(ws1, "已签到已签退", result["signed_both"], "E2EFDA")
    write_sheet(wb.create_sheet(), "已签到未签退", result["signed_no_checkout"], "FFF2CC")
    write_sheet(wb.create_sheet(), "未签到", result["not_signed"], "FCE4EC")
    write_sheet(wb.create_sheet(), "异常", result["anomalies"], "D9D9D9")

    stats = result["stats"]
    ws_sum = wb.create_sheet("汇总")
    ws_sum.append(["项目", "数量"])
    for k, v in [
        ("参加总人数", stats["attendee_total"]),
        ("已签到总人数", stats["signed_total"]),
        ("其中：已签到已签退", stats["signed_both"]),
        ("其中：已签到未签退", stats["signed_no_checkout"]),
        ("未签到人数", stats["not_signed_total"]),
        ("异常记录数", stats["anomaly_total"]),
    ]:
        ws_sum.append([k, v])
    for c in ws_sum[1]:
        c.font = Font(bold=True)

    fname = f"考勤比对_{mission_id}_{time.strftime('%Y%m%d_%H%M%S')}.xlsx"
    fpath = os.path.join(EXPORT_DIR, fname)
    wb.save(fpath)
    return FileResponse(fpath, filename=fname,
                        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ---------- 调试 ----------
@app.get("/api/debug/last")
def api_debug_last():
    return load_debug_last()


@app.get("/api/debug/ping")
def api_debug_ping():
    """无认证探活：确认网络与域名可达（需浏览器 UA，网站 WAF 拦截非浏览器请求）"""
    import requests
    try:
        r = requests.get("https://www.gdzyz.cn/", timeout=10, headers={
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                "(KHTML, like Gecko) Chrome/120.0 Safari/537.36 Edg/120.0"
            )
        })
        return {"ok": True, "http": r.status_code}
    except Exception as e:
        return {"ok": False, "error": str(e)}


# ---------- 静态文件 ----------
app.mount("/static", StaticFiles(directory=os.path.join(BASE_DIR, "static")), name="static")


@app.get("/")
def index():
    return FileResponse(os.path.join(BASE_DIR, "static", "index.html"))


if __name__ == "__main__":
    import uvicorn
    # 启动提示由 start.bat 展示，这里不再重复打印
    uvicorn.run(app, host="127.0.0.1", port=8000, log_level="warning")
